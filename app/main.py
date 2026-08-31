import io, json, os, re, secrets, tempfile
from datetime import datetime
from typing import Union
from urllib.parse import quote
from fastapi import FastAPI, Request, Form, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware

from app.models import *
from app.paths import TEMPLATES_DIR, STATIC_DIR
from app.auth import verify_pw, hash_pw, current_user, require
from app import calc, calc_sales, calc_purchasing, calc_arm, docgen

BASE_URL = os.environ.get("BASE_URL", "http://localhost:8000")

class JalurAsli:
    """Kembalikan alamat halaman yang sebenarnya diminta pengguna.

    Vercel meneruskan semua permintaan ke /api/index, sehingga aplikasi kehilangan
    alamat aslinya dan menjawab 404. Alamat asli dititipkan lewat parameter
    __vpath pada vercel.json; bila parameter itu tidak ada, awalan /api/index
    dipotong sebagai cadangan.
    """

    AWALAN = "/api/index"

    def __init__(self, app):
        self.app = app

    async def __call__(self, scope, receive, send):
        if scope.get("type") == "http":
            from urllib.parse import parse_qsl, urlencode
            scope = dict(scope)
            kueri = parse_qsl(scope.get("query_string", b"").decode(), keep_blank_values=True)
            jalur = None
            sisa = []
            for k, v in kueri:
                if k == "__vpath" and jalur is None:
                    jalur = v or "/"
                else:
                    sisa.append((k, v))
            if jalur:
                scope["path"] = jalur if jalur.startswith("/") else "/" + jalur
                scope["query_string"] = urlencode(sisa).encode()
            else:
                p = scope.get("path", "")
                if p == self.AWALAN or p.startswith(self.AWALAN + "/"):
                    scope["path"] = p[len(self.AWALAN):] or "/"
            scope["raw_path"] = scope["path"].encode()
        await self.app(scope, receive, send)


SIAP = {"sudah": False, "galat": None}


def siapkan_sekali():
    """Buat tabel dan data awal saat permintaan pertama, bukan saat impor.

    Dijalankan di sini (bukan di tingkat modul) supaya kesalahan konfigurasi
    database muncul sebagai halaman pesan yang terbaca, bukan fungsi yang gagal
    start tanpa keterangan.
    """
    if SIAP["sudah"]:
        return
    try:
        init_db()
        if os.environ.get("SKIP_SEED") != "1":
            import seed  # noqa: F401
        SIAP["galat"] = None
    except Exception as e:
        SIAP["galat"] = f"{type(e).__name__}: {e}"
        print("PENYIAPAN GAGAL:", SIAP["galat"], flush=True)
    SIAP["sudah"] = True


app = FastAPI(title="Aplikasi Pengajuan Insentif MFlash")
app.add_middleware(SessionMiddleware, secret_key=os.environ.get("SECRET", "ganti-secret-ini"))
app.add_middleware(JalurAsli)   # harus terluar: perbaiki alamat sebelum apa pun
if os.path.isdir(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
tpl = Jinja2Templates(directory=TEMPLATES_DIR)
tpl.env.globals.update(STATUS_LABELS=STATUS_LABELS, ROLE_LABELS=ROLE_LABELS,
                       TYPES=TYPES, BULAN=docgen.BULAN, rupiah=docgen.rupiah,
                       STATUS_FLOW=STATUS_FLOW, alur=alur)


def db_(): return SessionLocal()


@app.middleware("http")
async def penyiapan(request: Request, call_next):
    siapkan_sekali()
    if SIAP["galat"] and request.url.path not in ("/health", "/diagnosa"):
        galat = SIAP["galat"]
        if "OperationalError" in galat or "could not connect" in galat.lower():
            pesan = ("Aplikasi tidak bisa terhubung ke database. Periksa "
                     "DATABASE_URL di Environment Variables, lalu Redeploy. "
                     "Rinciannya ada di /diagnosa.")
        else:
            pesan = ("Penyiapan database gagal. Buka /diagnosa untuk melihat "
                     "pesan aslinya. Bila muncul nama kolom yang tidak dikenal, "
                     "cukup Redeploy sekali lagi agar struktur tabel diperbarui.")
        return tpl.TemplateResponse(
            request, "error.html",
            {"me": None, "kode": 500, "pesan": pesan}, status_code=500)
    return await call_next(request)


@app.get("/diagnosa")
def diagnosa():
    """Ringkasan kesehatan aplikasi. Tidak menampilkan nilai rahasia apa pun."""
    siapkan_sekali()
    url = os.environ.get("DATABASE_URL", "")

    # Periksa bentuk connection string tanpa membocorkan password
    from urllib.parse import urlparse
    pengguna = host = ""
    try:
        u = urlparse(url)
        pengguna, host = (u.username or ""), (u.hostname or "")
    except Exception:
        pass
    saran = []
    if url:
        if "pooler.supabase.com" in host and "." not in pengguna:
            saran.append(
                f"Nama pengguna '{pengguna}' salah untuk pooler Supabase. "
                f"Harus berbentuk 'postgres.<kode-project>', bukan 'postgres' saja. "
                f"Salin ulang dari tombol Connect di Supabase.")
        if ":5432/" in url and "pooler.supabase.com" in host:
            saran.append("Port 5432 adalah Session pooler. Untuk Vercel pakai "
                         "Transaction pooler port 6543.")
        if "[" in url or "]" in url:
            saran.append("Masih ada tanda kurung siku di connection string — "
                         "bagian [YOUR-PASSWORD] belum diganti.")
        if any(c in (url.split("@")[0] if "@" in url else "") for c in "@#?"):
            saran.append("Password kemungkinan mengandung karakter khusus. "
                         "Reset password di Supabase dan pilih yang hanya huruf dan angka.")
    else:
        saran.append("DATABASE_URL belum diisi di Environment Variables Vercel.")

    info = {
        "nama_pengguna": pengguna,
        "host": host,
        "saran": saran or ["Bentuk connection string terlihat benar."],
        "database_url_terisi": bool(url),
        "database": ("postgres" if "postgres" in url else
                     ("sqlite (DATABASE_URL belum diisi)" if not url else "lain")),
        "pooler": ("transaction 6543" if ":6543/" in url else
                   ("session 5432" if ":5432/" in url else "-")),
        "secret_terisi": bool(os.environ.get("SECRET")),
        "base_url": BASE_URL,
        "folder_template": os.path.isdir(TEMPLATES_DIR),
        "logo_tersedia": os.path.isfile(os.path.join(STATIC_DIR, "logo-mflash.png")),
        "berkas_rules": os.path.isfile(calc.RULES_PATH),
        "penyiapan_galat": SIAP["galat"],
    }
    try:
        db = db_()
        info["jumlah_cabang"] = db.query(Branch).count()
        db.close()
    except Exception as e:
        info["kesalahan_database"] = f"{type(e).__name__}: {e}"
    return info



def render(request, name, **ctx):
    ctx["me"] = current_user(request)
    return tpl.TemplateResponse(request, name, ctx)


@app.exception_handler(HTTPException)
def pesan_error(request: Request, exc: HTTPException):
    if exc.status_code == 401:
        return RedirectResponse("/login", 303)
    return tpl.TemplateResponse(request, "error.html",
                                {"me": current_user(request), "kode": exc.status_code,
                                 "pesan": exc.detail}, status_code=exc.status_code)


@app.get("/health")
def health():
    """Dipanggil ping otomatis tiap 10 menit: menjaga Render tetap bangun
    sekaligus menghitung aktivitas agar project Supabase tidak dijeda."""
    db = db_()
    try:
        n = db.query(Submission).count()
        return {"status": "ok", "pengajuan": n, "waktu": datetime.utcnow().isoformat()}
    finally:
        db.close()


# ------------------------------------------------- auth
@app.get("/login", response_class=HTMLResponse)
def login_form(request: Request, error: str = ""):
    return render(request, "login.html", error=error)


@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...)):
    db = db_()
    u = db.query(User).filter_by(username=username, active=True).first()
    if not u or not verify_pw(password, u.password_hash):
        return RedirectResponse("/login?error=Username+atau+password+salah", 303)
    request.session["uid"] = u.id
    return RedirectResponse("/", 303)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", 303)


# ------------------------------------------------- dashboard
PELIHAT_SEMUA = (ROLE_ARM, ROLE_CEO, ROLE_FINANCE, ROLE_ADMIN)


def boleh_lihat_nominal(u):
    """Store Leader tidak melihat nominal pengajuan di dashboard."""
    return u.role in PELIHAT_SEMUA


@app.get("/", response_class=HTMLResponse)
def home(request: Request, tab: str = "", cabang: int = 0, bulan: int = 0,
         tipe: str = "", tahun: int = 0, pesan: str = ""):
    u = current_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    db = db_()

    q = db.query(Submission)
    if u.role == ROLE_SL:
        q = q.filter(Submission.submitter_id == u.id)
    elif u.role == ROLE_ARM:
        # ARM melihat semua cabang, termasuk pengajuannya sendiri
        pass

    if cabang:
        q = q.filter(Submission.branch_id == cabang)
    if bulan:
        q = q.filter(Submission.period_month == bulan)
    if tahun:
        q = q.filter(Submission.period_year == tahun)
    if tipe:
        q = q.filter(Submission.type == tipe)

    semua = q.order_by(Submission.id.desc()).all()

    TAB = [("", "Semua"), (ST_DRAFT, "Draft"), (ST_WAIT_ARM, "Approval ARM"),
           (ST_WAIT_CEO, "Approval CEO"), (ST_WAIT_FIN, "Pencairan Finance"),
           (ST_DONE, "Done"), (ST_REJECTED, "Ditolak")]
    jumlah = {kode: len([x for x in semua if not kode or x.status == kode])
              for kode, _ in TAB}
    items = [x for x in semua if not tab or x.status == tab]
    inbox = [x for x in semua if ACTOR_OF_STATUS.get(x.status) == u.role]

    tautan = "".join(f"&{k}={v}" for k, v in
                     (("cabang", cabang), ("bulan", bulan), ("tahun", tahun),
                      ("tipe", tipe)) if v)
    # alamat untuk kembali ke tampilan yang sama sesudah menghapus
    url_kembali = quote(f"/?tab={tab}{tautan}", safe="")

    return render(request, "dashboard.html", items=items, inbox=inbox, tautan=tautan,
                  tab=tab, TAB=TAB, jumlah=jumlah,
                  f_cabang=cabang, f_bulan=bulan, f_tipe=tipe, f_tahun=tahun,
                  branches=db.query(Branch).order_by(Branch.name).all(),
                  tahun_ada=sorted({x.period_year for x in semua}, reverse=True),
                  lihat_nominal=boleh_lihat_nominal(u),
                  boleh_unduh=u.role in PELIHAT_SEMUA,
                  boleh_hapus=u.role == ROLE_ADMIN, pesan=pesan,
                  url_kembali=url_kembali)


# ------------------------------------------------- hapus pengajuan (admin)
@app.get("/pengajuan/{sid}/hapus", response_class=HTMLResponse)
def hapus_konfirmasi(request: Request, sid: int, kembali: str = "/"):
    require(request, [ROLE_ADMIN])
    db = db_()
    sub = db.query(Submission).get(sid)
    if not sub:
        raise HTTPException(404, "Pengajuan tidak ditemukan")
    return render(request, "hapus.html", sub=sub, kembali=kembali,
                  lampiran=db.query(Attachment).filter_by(submission_id=sid).count())


@app.post("/pengajuan/{sid}/hapus")
def hapus_pengajuan(request: Request, sid: int, kembali: str = Form("/")):
    require(request, [ROLE_ADMIN])
    db = db_()
    sub = db.query(Submission).get(sid)
    if not sub:
        raise HTTPException(404, "Pengajuan tidak ditemukan")
    kode = sub.code
    db.query(Attachment).filter_by(submission_id=sid).delete()
    db.query(Approval).filter_by(submission_id=sid).delete()
    db.delete(sub)
    db.commit()
    pemisah = "&" if "?" in kembali else "?"
    return RedirectResponse(f"{kembali}{pemisah}pesan=Pengajuan+{kode}+dihapus", 303)


# ------------------------------------------------- kelola akun
MIN_PASSWORD = 8


def _admin_aktif_lain(db, kecuali_id):
    """Jumlah admin aktif selain akun tertentu — penjaga agar tidak terkunci."""
    return (db.query(User)
              .filter(User.role == ROLE_ADMIN, User.active == True,  # noqa: E712
                      User.id != kecuali_id)
              .count())


@app.get("/master/akun", response_class=HTMLResponse)
def master_akun(request: Request, edit: int = 0, pesan: str = "", galat: str = ""):
    require(request, [ROLE_ADMIN])
    db = db_()
    return render(request, "master_akun.html",
                  items=db.query(User).order_by(User.role, User.username).all(),
                  branches=db.query(Branch).filter_by(active=True)
                             .order_by(Branch.name).all(),
                  edit=db.query(User).get(edit) if edit else None,
                  pesan=pesan, galat=galat)


@app.post("/master/akun")
def master_akun_simpan(request: Request, id: int = Form(0), username: str = Form(...),
                       full_name: str = Form(...), position: str = Form(""),
                       role: str = Form(...), branch_id: int = Form(0),
                       password: str = Form("")):
    me = require(request, [ROLE_ADMIN])
    db = db_()
    username = username.strip().lower()
    if not username or not full_name.strip():
        return RedirectResponse("/master/akun?galat=Username+dan+nama+wajib+diisi", 303)
    if role not in ROLE_LABELS:
        return RedirectResponse("/master/akun?galat=Peran+tidak+dikenali", 303)
    kembar = db.query(User).filter(User.username == username, User.id != id).first()
    if kembar:
        return RedirectResponse(f"/master/akun?galat=Username+{username}+sudah+dipakai", 303)

    u = db.query(User).get(id) if id else User()
    if not id and len(password) < MIN_PASSWORD:
        return RedirectResponse(
            f"/master/akun?galat=Password+minimal+{MIN_PASSWORD}+karakter", 303)
    # jangan sampai admin terakhir kehilangan perannya sendiri
    if id and u and u.role == ROLE_ADMIN and role != ROLE_ADMIN \
            and _admin_aktif_lain(db, u.id) == 0:
        return RedirectResponse(
            "/master/akun?galat=Ini+administrator+terakhir,+perannya+tidak+bisa+diubah", 303)

    u.username, u.full_name = username, full_name.strip()
    u.position, u.role = position.strip(), role
    u.branch_id = branch_id or None
    if password:
        if len(password) < MIN_PASSWORD:
            return RedirectResponse(
                f"/master/akun?galat=Password+minimal+{MIN_PASSWORD}+karakter", 303)
        u.password_hash = hash_pw(password)
    if not id:
        u.active = True
        db.add(u)
    db.commit()
    return RedirectResponse("/master/akun?pesan=Akun+tersimpan", 303)


@app.post("/master/akun/{uid}/aktif")
def master_akun_aktif(request: Request, uid: int):
    me = require(request, [ROLE_ADMIN])
    db = db_()
    u = db.query(User).get(uid)
    if not u:
        return RedirectResponse("/master/akun?galat=Akun+tidak+ditemukan", 303)
    if u.id == me.id:
        return RedirectResponse(
            "/master/akun?galat=Anda+tidak+bisa+menonaktifkan+akun+sendiri", 303)
    if u.active and u.role == ROLE_ADMIN and _admin_aktif_lain(db, u.id) == 0:
        return RedirectResponse(
            "/master/akun?galat=Ini+administrator+terakhir+yang+aktif", 303)
    u.active = not u.active
    db.commit()
    return RedirectResponse("/master/akun?pesan=Status+akun+diperbarui", 303)


@app.get("/akun/password", response_class=HTMLResponse)
def form_password(request: Request, pesan: str = "", galat: str = ""):
    require(request)
    return render(request, "ganti_password.html", pesan=pesan, galat=galat,
                  minimal=MIN_PASSWORD)


@app.post("/akun/password")
def ganti_password(request: Request, lama: str = Form(...), baru: str = Form(...),
                   ulangi: str = Form(...)):
    me = require(request)
    db = db_()
    u = db.query(User).get(me.id)
    if not verify_pw(lama, u.password_hash):
        return RedirectResponse("/akun/password?galat=Password+lama+salah", 303)
    if len(baru) < MIN_PASSWORD:
        return RedirectResponse(
            f"/akun/password?galat=Password+baru+minimal+{MIN_PASSWORD}+karakter", 303)
    if baru != ulangi:
        return RedirectResponse("/akun/password?galat=Ulangan+password+tidak+sama", 303)
    if baru == lama:
        return RedirectResponse("/akun/password?galat=Password+baru+sama+dengan+yang+lama", 303)
    u.password_hash = hash_pw(baru)
    db.commit()
    return RedirectResponse("/akun/password?pesan=Password+berhasil+diganti", 303)


# ------------------------------------------------- master cabang
@app.get("/master/cabang", response_class=HTMLResponse)
def master_cabang(request: Request, edit: int = 0, pesan: str = ""):
    require(request, [ROLE_ADMIN, ROLE_ARM])
    db = db_()
    return render(request, "master_cabang.html",
                  items=db.query(Branch).order_by(Branch.name).all(),
                  edit=db.query(Branch).get(edit) if edit else None, pesan=pesan)


@app.post("/master/cabang")
def master_cabang_simpan(request: Request, id: int = Form(0), name: str = Form(...),
                         display_name: str = Form(""), address: str = Form(""),
                         city: str = Form("Jakarta"), hitung_arm: str = Form("")):
    require(request, [ROLE_ADMIN, ROLE_ARM])
    db = db_()
    name = name.strip()
    if not name:
        return RedirectResponse("/master/cabang?pesan=Nama+cabang+wajib+diisi", 303)
    kembar = db.query(Branch).filter(Branch.name == name, Branch.id != id).first()
    if kembar:
        return RedirectResponse(f"/master/cabang?pesan=Cabang+{name}+sudah+ada", 303)
    b = db.query(Branch).get(id) if id else Branch()
    b.name = name
    b.display_name = display_name.strip() or f"MFlash \u2013 {name}"
    b.address = address.strip()
    b.city = city.strip() or "Jakarta"
    b.hitung_arm = bool(hitung_arm)
    if not id:
        b.active = True
        db.add(b)
    db.commit()
    return RedirectResponse("/master/cabang?pesan=Data+cabang+tersimpan", 303)


@app.post("/master/cabang/{bid}/aktif")
def master_cabang_aktif(request: Request, bid: int):
    require(request, [ROLE_ADMIN, ROLE_ARM])
    db = db_()
    b = db.query(Branch).get(bid)
    if b:
        b.active = not b.active
        db.commit()
    return RedirectResponse("/master/cabang", 303)


# ------------------------------------------------- master sales
@app.get("/master/sales", response_class=HTMLResponse)
def master_sales(request: Request, edit: int = 0, pesan: str = "", galat: str = ""):
    require(request, [ROLE_ADMIN, ROLE_ARM])
    db = db_()
    return render(request, "master_sales.html",
                  items=db.query(Sales).order_by(Sales.branch_id, Sales.name).all(),
                  branches=db.query(Branch).filter_by(active=True).order_by(Branch.name).all(),
                  edit=db.query(Sales).get(edit) if edit else None,
                  pesan=pesan, galat=galat)


@app.post("/master/sales")
def master_sales_simpan(request: Request, id: int = Form(0), name: str = Form(...),
                        branch_id: int = Form(...), status_karyawan: str = Form("TEAM INTI"),
                        aliases: str = Form("")):
    require(request, [ROLE_ADMIN, ROLE_ARM])
    db = db_()
    name = name.strip()
    if not name:
        return RedirectResponse("/master/sales?pesan=Nama+sales+wajib+diisi", 303)
    sl = db.query(Sales).get(id) if id else Sales()
    sl.name, sl.branch_id = name, branch_id
    sl.status_karyawan = status_karyawan.strip() or "TEAM INTI"
    sl.aliases = aliases.strip()
    if not id:
        sl.active = True
        db.add(sl)
    db.commit()
    return RedirectResponse("/master/sales?pesan=Data+sales+tersimpan", 303)


@app.post("/master/sales/impor")
async def master_sales_impor(request: Request, berkas: UploadFile = File(...),
                             branch_id: int = Form(0)):
    """Impor daftar sales dari Excel: kolom Nama, Cabang, Status Karyawan, Ejaan Lain."""
    require(request, [ROLE_ADMIN, ROLE_ARM])
    db = db_()
    if not getattr(berkas, "filename", ""):
        return RedirectResponse("/master/sales?galat=Pilih+file+dulu", 303)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as t:
        t.write(await berkas.read())
        path = t.name

    def _k(x):
        return re.sub(r"[^a-z0-9]", "", str(x or "").lower())

    ALIAS = {"nama": "nama", "namasales": "nama", "nama sales": "nama",
             "cabang": "cabang", "status": "status",
             "statuskaryawan": "status", "ejaanlain": "alias",
             "alias": "alias", "ejaanlainpisahkankoma": "alias"}
    baru = ubah = lewat = 0
    tak_ada_cabang = set()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        header_i = None
        for i, r in enumerate(rows[:15]):
            kunci = {ALIAS.get(_k(c)) for c in r if c is not None}
            if "nama" in kunci:
                header_i = i
                break
        if header_i is None:
            raise ValueError("Kolom Nama tidak ditemukan pada file")
        header = [ALIAS.get(_k(c)) for c in rows[header_i]]

        cabang_peta = {_k(b.name): b for b in db.query(Branch).all()}
        for r in rows[header_i + 1:]:
            d = {h: v for h, v in zip(header, r) if h}
            nama = str(d.get("nama") or "").strip()
            if not nama:
                continue
            nm_cabang = str(d.get("cabang") or "").strip()
            b = cabang_peta.get(_k(nm_cabang)) if nm_cabang else None
            if not b and branch_id:
                b = db.query(Branch).get(branch_id)
            if not b:
                tak_ada_cabang.add(nm_cabang or "(kosong)")
                lewat += 1
                continue
            sl = db.query(Sales).filter_by(name=nama, branch_id=b.id).first()
            status = str(d.get("status") or "TEAM INTI").strip() or "TEAM INTI"
            alias = str(d.get("alias") or "").strip()
            if sl:
                sl.status_karyawan, sl.aliases, sl.active = status, alias, True
                ubah += 1
            else:
                db.add(Sales(name=nama, branch_id=b.id, status_karyawan=status,
                             aliases=alias, active=True))
                baru += 1
        db.commit()
        pesan = f"Impor selesai: {baru} sales baru, {ubah} diperbarui"
        if lewat:
            pesan += (f", {lewat} dilewati karena cabangnya tidak dikenali "
                      f"({', '.join(sorted(tak_ada_cabang)[:5])})")
    except Exception as e:
        pesan = f"Impor gagal: {e}"
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass
    return RedirectResponse(f"/master/sales?pesan={pesan.replace(' ', '+')}", 303)


@app.get("/master/sales/contoh.xlsx")
def master_sales_contoh(request: Request):
    """Berkas contoh berisi kolom yang dikenali beserta data cabang yang ada."""
    require(request, [ROLE_ADMIN, ROLE_ARM])
    db = db_()
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Nama", "Cabang", "Status Karyawan", "Ejaan Lain"])
    contoh = db.query(Branch).filter_by(active=True).order_by(Branch.name).first()
    ws.append(["ALVANDI KLENDER", contoh.name if contoh else "Klender",
               "TEAM INTI", ""])
    ws.append(["ARIF FIKRI KLENDER", contoh.name if contoh else "Klender",
               "TEAM INTI", "Arif Fikri"])
    for lebar, kolom in zip((28, 18, 18, 26), "ABCD"):
        ws.column_dimensions[kolom].width = lebar
    out = os.path.join(tempfile.gettempdir(), "Contoh Impor Master Sales.xlsx")
    wb.save(out)
    return FileResponse(out, filename="Contoh Impor Master Sales.xlsx")


@app.post("/master/sales/{sid}/aktif")
def master_sales_aktif(request: Request, sid: int):
    require(request, [ROLE_ADMIN, ROLE_ARM])
    db = db_()
    sl = db.query(Sales).get(sid)
    if sl:
        sl.active = not sl.active
        db.commit()
    return RedirectResponse("/master/sales", 303)


# ------------------------------------------------- master supplier
@app.get("/master/supplier", response_class=HTMLResponse)
def master_supplier(request: Request, edit: int = 0, cari: str = "", pesan: str = ""):
    require(request, [ROLE_ADMIN, ROLE_ARM])
    db = db_()
    q = db.query(Supplier)
    if cari:
        q = q.filter(Supplier.name.ilike(f"%{cari}%"))
    items = q.order_by(Supplier.name).limit(300).all()
    return render(request, "master_supplier.html", items=items, cari=cari, pesan=pesan,
                  edit=db.query(Supplier).get(edit) if edit else None,
                  jml=db.query(Supplier).count(),
                  jml_target=db.query(Supplier)
                               .filter_by(kategori="TERTARGET").count())


@app.post("/master/supplier")
def master_supplier_simpan(request: Request, id: int = Form(0), name: str = Form(...),
                           kategori: str = Form("NON TERTARGET")):
    require(request, [ROLE_ADMIN, ROLE_ARM])
    db = db_()
    name = name.strip()
    if not name:
        return RedirectResponse("/master/supplier?pesan=Nama+pemasok+wajib+diisi", 303)
    kembar = db.query(Supplier).filter(Supplier.name == name, Supplier.id != id).first()
    if kembar:
        return RedirectResponse("/master/supplier?pesan=Pemasok+sudah+terdaftar", 303)
    sp = db.query(Supplier).get(id) if id else Supplier()
    sp.name, sp.kategori = name, kategori
    if not id:
        db.add(sp)
    db.commit()
    return RedirectResponse("/master/supplier?pesan=Data+pemasok+tersimpan", 303)


@app.post("/master/supplier/impor")
async def master_supplier_impor(request: Request, berkas: UploadFile = File(...)):
    """Impor daftar kategori supplier dari file Excel (kolom Pemasok + kategori)."""
    require(request, [ROLE_ADMIN, ROLE_ARM])
    db = db_()
    if not getattr(berkas, "filename", ""):
        return RedirectResponse("/master/supplier?pesan=Pilih+file+dulu", 303)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as t:
        t.write(await berkas.read())
        path = t.name
    baru = ubah = 0
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = None
        for nm in wb.sheetnames:
            if "supplier" in nm.lower() or "pemasok" in nm.lower():
                ws = wb[nm]
                break
        ws = ws or wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
            nama = str(row[0] or "").strip()
            kat = str(row[1] or "").strip().upper()
            if not nama or kat not in ("TERTARGET", "NON TERTARGET"):
                continue
            sp = db.query(Supplier).filter_by(name=nama).first()
            if sp:
                if sp.kategori != kat:
                    sp.kategori = kat
                    ubah += 1
            else:
                db.add(Supplier(name=nama, kategori=kat))
                baru += 1
        db.commit()
        pesan = f"Impor selesai: {baru} pemasok baru, {ubah} kategori diperbarui"
    except Exception as e:
        pesan = f"Impor gagal: {e}"
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass
    return RedirectResponse(f"/master/supplier?pesan={pesan.replace(' ', '+')}", 303)


# ------------------------------------------------- pengajuan baru
def _jenis_boleh(u):
    """Jenis pengajuan yang boleh dibuat oleh peran ini."""
    return {k: v for k, v in TYPES.items()
            if u.role in PENGAJU_JENIS.get(k, (ROLE_SL, ROLE_ADMIN))}


@app.get("/pengajuan/baru", response_class=HTMLResponse)
def new_form(request: Request):
    u = require(request, [ROLE_SL, ROLE_ARM, ROLE_ADMIN])
    db = db_()
    jenis = _jenis_boleh(u)
    if not jenis:
        raise HTTPException(403, "Peran Anda tidak dapat membuat pengajuan")
    return render(request, "new.html", jenis_tersedia=jenis,
                  branches=db.query(Branch).filter_by(active=True)
                             .order_by(Branch.name).all())


@app.post("/pengajuan/baru")
async def new_submit(request: Request,
                     jenis: str = Form(...), nama: str = Form(...),
                     bulan: int = Form(...), tahun: int = Form(...),
                     laba_ditahan_pct: float = Form(7),
                     goal_nama: list[str] = Form([]), goal_pct: list[str] = Form([]),
                     blok_tipe: list[str] = Form([]),
                     b_asal: list[int] = Form([]), b_tujuan: list[int] = Form([]),
                     mut_bulan: list[int] = Form([]), mut_tahun: list[int] = Form([]),
                     excel_asal: list[Union[UploadFile, str]] = File([]),
                     excel_tujuan: list[Union[UploadFile, str]] = File([]),
                     s_branch: int = Form(0), p_branch: int = Form(0),
                     f_pelanggan: Union[UploadFile, str] = File(None),
                     f_faktur: Union[UploadFile, str] = File(None),
                     f_accurate: Union[UploadFile, str] = File(None),
                     f_arm: Union[UploadFile, str] = File(None),
                     prioritas_pct: str = Form(""),
                     arm_cabang: list[int] = Form([]),
                     shift_sl: str = Form("")):
    u = require(request, [ROLE_SL, ROLE_ARM, ROLE_ADMIN])
    db = db_()
    if u.role not in PENGAJU_JENIS.get(jenis, (ROLE_SL, ROLE_ADMIN)):
        raise HTTPException(403, "Peran Anda tidak dapat mengajukan jenis ini")
    if jenis not in ("sales_team", "purchasing", "profit_arm") and not blok_tipe:
        raise HTTPException(400, "Tambahkan minimal satu cabang.")

    if jenis == "sales_team":
        cabang_utama = s_branch
    elif jenis == "purchasing":
        cabang_utama = p_branch
    elif jenis == "profit_arm":
        pusat = db.query(Branch).filter_by(name="Pusat").first()
        cabang_utama = pusat.id if pusat else (b_asal[0] if b_asal else None)
    else:
        # Kop surat dan cabang pengajuan mengikuti cabang tujuan bila Store
        # Leader sedang dalam masa rotasi; cabang asal hanya sumber angka.
        cabang_utama = b_asal[0]
        for i, t in enumerate(blok_tipe):
            if t == "rotasi" and i < len(b_tujuan) and b_tujuan[i]:
                cabang_utama = b_tujuan[i]
                break
    code = f"INS/{datetime.now():%Y%m}/{secrets.token_hex(3).upper()}"
    sub = Submission(code=code, type=jenis, branch_id=cabang_utama, submitter_id=u.id,
                     submitter_name=(nama or "").strip() or u.full_name,
                     period_month=bulan, period_year=tahun, status=ST_DRAFT)
    db.add(sub); db.commit()

    async def simpan(f, label):
        """Tulis file ke temp untuk dihitung, sekaligus arsipkan ke database."""
        if not getattr(f, "filename", ""):
            raise calc.CalcError(f"File Laporan Laba/Rugi cabang {label} belum dipilih.")
        raw = await f.read()
        db.add(Attachment(submission_id=sub.id, kind="excel",
                          filename=f"{label} - {f.filename}",
                          mime="application/vnd.openxmlformats-officedocument"
                               ".spreadsheetml.sheet", blob=raw))
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as t:
            t.write(raw)
            return t.name

    nama_cabang = {b.id: b.name for b in db.query(Branch).all()}

    prioritas = (float(prioritas_pct.replace(",", "."))
                 if prioritas_pct.strip() else None)
    goals = []
    for i, nm in enumerate(goal_nama):
        nilai = goal_pct[i].strip() if i < len(goal_pct) else ""
        goals.append({"nama": nm.strip(),
                      "pencapaian": float(nilai.replace(",", ".")) if nilai else None})

    blok_list, temps = [], []
    try:
        if jenis == "sales_team":
            p1 = await simpan(f_pelanggan, "Data Pelanggan")
            p2 = await simpan(f_faktur, "Rincian Faktur Penjualan")
            temps += [p1, p2]
            def daftar_sales_cabang():
                return [{"nama": x.name, "status": x.status_karyawan,
                         "alias": x.aliases}
                        for x in db.query(Sales)
                                   .filter_by(branch_id=cabang_utama, active=True)
                                   .order_by(Sales.name).all()]

            def daftarkan_baru(nama_baru):
                """Nama penjual yang muncul di faktur tapi belum ada di Master
                Sales langsung didaftarkan, supaya insentifnya tidak hilang."""
                for n in nama_baru:
                    db.add(Sales(branch_id=cabang_utama, name=n,
                                 status_karyawan="Team Inti", active=True))
                db.commit()
                return daftar_sales_cabang()

            daftar = daftar_sales_cabang()
            hasil = calc_sales.hitung_sales(p1, p2, bulan, tahun, daftar,
                                            daftarkan_baru=daftarkan_baru)
            sub.total_amount = hasil["total"]
            sub.data_json = json.dumps(hasil, default=str)
            catatan = []
            if hasil.get("nama_baru_didaftarkan"):
                nb = hasil["nama_baru_didaftarkan"]
                catatan.append(f"{len(nb)} nama penjual belum ada di Master Sales dan "
                               "otomatis ditambahkan: " + ", ".join(nb[:15])
                               + ("…" if len(nb) > 15 else "")
                               + ". Periksa di menu Master Sales bila ada yang salah "
                                 "ketik atau bukan sales.")
            if hasil["nama_tak_dikenal"]:
                catatan.append("Nama penjual berikut ada di data tapi belum terdaftar "
                               "di Master Sales, jadi transaksinya tidak dihitung: "
                               + ", ".join(hasil["nama_tak_dikenal"][:15]))
            sub.note = " ".join(catatan) or None
            db.commit()
            return RedirectResponse(f"/pengajuan/{sub.id}", 303)

        if jenis == "profit_arm":
            p1 = await simpan(f_arm, "Laporan Keuangan MGI")
            temps.append(p1)
            # Pengaju boleh mencentang sendiri cabang mana yang dihitung.
            # Nilai 0 adalah penanda bahwa daftar centang ikut terkirim; tanpa
            # penanda itu kita pakai pengaturan Master Cabang.
            dipilih = {i for i in arm_cabang if i}
            pakai_pilihan = 0 in arm_cabang
            daftar = [{"nama": b.name,
                       "ikut": (b.id in dipilih) if pakai_pilihan
                               else bool(b.hitung_arm)}
                      for b in db.query(Branch).filter_by(active=True)
                                 .order_by(Branch.name).all()]
            if pakai_pilihan and not dipilih:
                raise calc.CalcError("Pilih minimal satu cabang yang dihitung.")
            hasil = calc_arm.hitung_arm(p1, bulan, tahun, daftar,
                                        laba_ditahan_pct, prioritas, goals)
            sub.total_amount = hasil["total"]
            sub.data_json = json.dumps(hasil, default=str)
            if hasil["cabang_tak_ketemu"]:
                sub.note = ("Sheet untuk cabang berikut tidak ditemukan di file: "
                            + ", ".join(hasil["cabang_tak_ketemu"]))
            db.commit()
            return RedirectResponse(f"/pengajuan/{sub.id}", 303)

        if jenis == "purchasing":
            p1 = await simpan(f_accurate, "Data Accurate")
            temps.append(p1)
            master = {x.name: x.kategori for x in db.query(Supplier).all()}
            hasil = calc_purchasing.hitung_purchasing(p1, bulan, tahun, master)
            sub.total_amount = hasil["total"]
            sub.data_json = json.dumps(hasil, default=str)
            if hasil["supplier_tak_terdaftar"]:
                sub.note = (f"{len(hasil['supplier_tak_terdaftar'])} pemasok belum "
                            f"terdaftar di Master Supplier senilai "
                            f"{docgen.rupiah(hasil['nilai_tak_terdaftar'])} "
                            f"— dikeluarkan dari perhitungan.")
            db.commit()
            return RedirectResponse(f"/pengajuan/{sub.id}", 303)

        for i, tipe in enumerate(blok_tipe):
            asal = nama_cabang.get(b_asal[i], "-")
            p1 = await simpan(excel_asal[i], asal)
            temps.append(p1)
            blok = {"tipe": tipe, "asal": {"cabang": asal, "path": p1}}
            if tipe == "rotasi":
                tujuan = nama_cabang.get(b_tujuan[i], "-")
                p2 = await simpan(excel_tujuan[i], tujuan)
                temps.append(p2)
                blok.update({"tujuan": {"cabang": tujuan, "path": p2},
                             "mutasi_bulan": mut_bulan[i], "mutasi_tahun": mut_tahun[i]})
            blok_list.append(blok)

        hasil = calc.hitung_pengajuan(blok_list, bulan, tahun, laba_ditahan_pct,
                                      goals, shift_sl)
        sub.total_amount = hasil["total"]
        sub.data_json = json.dumps(hasil, default=str)
        if hasil["catatan"]:
            sub.note = " | ".join(hasil["catatan"])
    except calc.CalcError as e:
        sub.note = str(e)
    except Exception as e:
        sub.note = f"Perhitungan otomatis gagal: {e}"
    finally:
        for t in temps:
            try:
                os.unlink(t)
            except OSError:
                pass
    db.commit()
    return RedirectResponse(f"/pengajuan/{sub.id}", 303)


# ------------------------------------------------- detail & aksi
@app.get("/pengajuan/{sid}", response_class=HTMLResponse)
def detail(request: Request, sid: int):
    u = require(request)
    db = db_()
    s = db.query(Submission).get(sid)
    if not s:
        raise HTTPException(404, "Pengajuan tidak ditemukan")
    if u.role == ROLE_SL and s.submitter_id != u.id:
        raise HTTPException(403, "Bukan pengajuan Anda")
    hasil = json.loads(s.data_json or "{}")
    can_act = ACTOR_OF_STATUS.get(s.status) == u.role
    return render(request, "detail.html", s=s, hasil=hasil, can_act=can_act)


@app.post("/pengajuan/{sid}/submit")
def do_submit(request: Request, sid: int):
    u = require(request, [ROLE_SL, ROLE_ADMIN])
    db = db_()
    s = db.query(Submission).get(sid)
    if s.status != ST_DRAFT:
        raise HTTPException(400, "Pengajuan sudah dikirim")
    s.status = alur(s.type)[0]
    db.add(Approval(submission_id=s.id, user_id=u.id, role=u.role, action="submit",
                    qr_token=secrets.token_urlsafe(12)))
    db.commit()
    return RedirectResponse(f"/pengajuan/{sid}", 303)


@app.post("/pengajuan/{sid}/aksi")
def do_action(request: Request, sid: int, aksi: str = Form(...), catatan: str = Form("")):
    u = require(request)
    db = db_()
    s = db.query(Submission).get(sid)
    if ACTOR_OF_STATUS.get(s.status) != u.role:
        raise HTTPException(403, "Bukan giliran role Anda untuk memproses")
    token = secrets.token_urlsafe(12)
    if aksi == "tolak":
        s.status = ST_REJECTED
        act = "reject"
    else:
        act = "approve"
        urut = alur(s.type)
        s.status = urut[urut.index(s.status) + 1]
    db.add(Approval(submission_id=s.id, user_id=u.id, role=u.role, action=act,
                    note=catatan, qr_token=token))
    db.commit()
    return RedirectResponse(f"/pengajuan/{sid}", 303)


# ------------------------------------------------- dokumen word
MIME_DOCX = ("application/vnd.openxmlformats-officedocument"
             ".wordprocessingml.document")


@app.get("/pengajuan/{sid}/dokumen.docx")
@app.get("/pengajuan/{sid}/docx")          # alamat lama tetap dilayani
def download_docx(request: Request, sid: int):
    u = require(request)
    db = db_()
    s = db.query(Submission).get(sid)
    if not s:
        raise HTTPException(404, "Pengajuan tidak ditemukan")
    if u.role == ROLE_SL and s.submitter_id != u.id:
        raise HTTPException(403, "Bukan pengajuan Anda")

    hasil = json.loads(s.data_json or "{}")
    out = os.path.join(tempfile.gettempdir(), s.code.replace("/", "_") + ".docx")
    docgen.buat_dokumen(s, hasil, s.approvals, BASE_URL, out)

    nama = (f"{TYPES[s.type]['label']} - {s.branch.name} - "
            f"{docgen.BULAN[s.period_month]} {s.period_year}.docx")
    # Kirim nama berkas dalam dua bentuk: teks biasa sebagai cadangan, dan
    # bentuk terkode. Tanpa cadangan, sebagian browser memakai potongan alamat
    # sebagai nama berkas sehingga .docx dikira arsip dan dibuka jadi folder.
    polos = "".join(c if (c.isalnum() or c in " .-_") else "-" for c in nama)
    disposisi = (f'attachment; filename="{polos}"; '
                 f"filename*=UTF-8''{quote(nama)}")
    return FileResponse(out, media_type=MIME_DOCX,
                        headers={"Content-Disposition": disposisi,
                                 "X-Content-Type-Options": "nosniff"})


@app.post("/pengajuan/unduh-massal")
def unduh_massal(request: Request, sid: list[int] = Form([])):
    """Kumpulkan dokumen Word beberapa pengajuan menjadi satu berkas .zip."""
    u = require(request, [ROLE_ARM, ROLE_CEO, ROLE_FINANCE, ROLE_ADMIN])
    db = db_()
    if not sid:
        raise HTTPException(400, "Tidak ada pengajuan yang dipilih")

    daftar = (db.query(Submission).filter(Submission.id.in_(sid))
                .order_by(Submission.branch_id, Submission.id).all())
    if not daftar:
        raise HTTPException(404, "Pengajuan tidak ditemukan")

    import zipfile
    zip_path = os.path.join(tempfile.gettempdir(),
                            f"Pengajuan Insentif {datetime.now():%Y%m%d-%H%M%S}.zip")
    dipakai = set()
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for s in daftar:
            hasil = json.loads(s.data_json or "{}")
            tmp = os.path.join(tempfile.gettempdir(),
                               s.code.replace("/", "_") + ".docx")
            docgen.buat_dokumen(s, hasil, s.approvals, BASE_URL, tmp)
            nama = (f"{docgen.BULAN[s.period_month]} {s.period_year} - "
                    f"{s.branch.name if s.branch else '-'} - "
                    f"{TYPES[s.type]['label']}.docx")
            nama = "".join(c if (c.isalnum() or c in " .-_") else "-" for c in nama)
            asli, n = nama, 2
            while nama in dipakai:
                nama = asli.replace(".docx", f" ({n}).docx")
                n += 1
            dipakai.add(nama)
            z.write(tmp, nama)

    disposisi = (f'attachment; filename="{os.path.basename(zip_path)}"; '
                 f"filename*=UTF-8''{quote(os.path.basename(zip_path))}")
    return FileResponse(zip_path, media_type="application/zip",
                        headers={"Content-Disposition": disposisi})


@app.get("/verify/{token}", response_class=HTMLResponse)
def verify(request: Request, token: str):
    db = db_()
    a = db.query(Approval).filter_by(qr_token=token).first()
    return render(request, "verify.html", a=a)


@app.get("/lampiran/{aid}")
def lampiran(request: Request, aid: int):
    require(request)
    db = db_()
    at = db.query(Attachment).get(aid)
    if not at:
        raise HTTPException(404, "Lampiran tidak ditemukan")
    return StreamingResponse(io.BytesIO(at.blob), media_type=at.mime or "application/octet-stream",
        headers={"Content-Disposition": f'inline; filename="{at.filename}"'})
