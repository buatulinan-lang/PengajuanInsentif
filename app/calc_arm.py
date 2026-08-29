"""
Insentif Profit ARM.

Masukan: satu file Laporan Keuangan MGI berisi seluruh cabang, tiap cabang
menempati satu sheet ("LR MF Klender (AR)", "LR MF Ceger (SE)", dan seterusnya).

Mekanisme:
  1. Laba bersih tiap cabang dihitung sama seperti Insentif Profit Store Leader
     (Total Laba Bersih Setelah Prepaid dikurangi Laba Ditahan).
  2. Cabang yang ditandai tidak ikut hitung ARM dikecualikan.
  3. Total laba bersih seluruh cabang -> baris matriks (dibulatkan ke bawah).
  4. Nominal dikali pengali menurut Pencapaian Target Prioritas.
  5. Hasilnya dikurangi potongan Goal Utama / Pendamping, aturannya sama
     dengan Insentif Profit Store Leader.
Angka-angkanya ada di data/rules.json bagian "profit_arm".
"""
import json
import re

from openpyxl import load_workbook

from app.calc import (RULES_PATH, CalcError, hitung_profit_sl,
                      hitung_pengurang_goal, BULAN_XLS)


def _cocok(nama_sheet, nama_cabang):
    """Sheet 'LR MF Jatiwaringin (WB)' cocok dengan cabang 'Jatiwaringin'."""
    bersih = re.sub(r"[^a-z0-9]", "", nama_sheet.lower())
    return re.sub(r"[^a-z0-9]", "", nama_cabang.lower()) in bersih


def hitung_arm(path, bulan, tahun=None, cabang_list=None,
               laba_ditahan_pct=None, pencapaian_prioritas=None, goals=None):
    """
    cabang_list: [{"nama": str, "ikut": bool}] dari master cabang.
    """
    rules = json.load(open(RULES_PATH))["profit_arm"]
    if laba_ditahan_pct is None:
        laba_ditahan_pct = rules.get("laba_ditahan_pct_default", 7)

    # Buka berkas sekali saja; membuka ulang tiap cabang membuat file besar
    # (17 sheet) butuh lebih dari satu menit.
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = wb.sheetnames
    isi = {}
    try:
        perlu = {}
        for c in (cabang_list or []):
            sheet = next((s for s in sheets if _cocok(s, c["nama"])), None)
            if sheet:
                perlu[c["nama"]] = sheet
        for nama_c, sheet in perlu.items():
            isi[sheet] = [r for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()

    baris, tak_ketemu = [], []
    total = 0.0
    for c in (cabang_list or []):
        sheet = next((s for s in sheets if _cocok(s, c["nama"])), None)
        if not sheet:
            tak_ketemu.append(c["nama"])
            continue
        try:
            h = hitung_profit_sl(path, bulan, laba_ditahan_pct,
                                 rows=isi.get(sheet))
        except CalcError as e:
            baris.append({"cabang": c["nama"], "sheet": sheet, "ikut": False,
                          "galat": str(e), "omzet": 0, "laba_setelah_prepaid": 0,
                          "laba_ditahan": 0, "laba_bersih": 0})
            continue
        ikut = bool(c.get("ikut", True))
        if ikut:
            total += h["laba_bersih"]
        baris.append({"cabang": c["nama"], "sheet": sheet, "ikut": ikut,
                      "omzet": h["omzet"],
                      "laba_setelah_prepaid": h["laba_setelah_prepaid"],
                      "laba_ditahan": h["laba_ditahan"],
                      "laba_bersih": h["laba_bersih"], "galat": None})

    if not baris:
        raise CalcError("Tidak ada sheet cabang yang cocok di file laporan "
                        "keuangan. Pastikan file yang diunggah benar.")

    # baris matriks: turun ke ambang terdekat di bawah total laba bersih
    nominal, baris_matriks = 0, None
    for m in rules["matriks"]:
        if total >= m["laba_min"]:
            nominal, baris_matriks = m["insentif"], m["laba_min"]

    # pengali menurut pencapaian target prioritas
    pengali, label_pengali = 0, "-"
    if pencapaian_prioritas is not None:
        for t in rules["pengali_target_prioritas"]:
            if pencapaian_prioritas >= t["pencapaian_min"]:
                pengali, label_pengali = t["pengali_pct"], t["label"]
                break
    else:
        pengali, label_pengali = 100, "belum diisi (dianggap 100%)"

    setelah_pengali = round(nominal * pengali / 100)

    g = hitung_pengurang_goal(goals)
    potongan = round(setelah_pengali * g["potongan_pct"] / 100)
    for p in g.get("rincian_potongan", []):
        p["nilai"] = round(setelah_pengali * p["pct"] / 100)

    return {
        "jenis": "profit_arm", "bulan": bulan, "tahun": tahun,
        "bulan_nama": BULAN_XLS[bulan - 1].title(),
        "baris": baris,
        "cabang_tak_ketemu": tak_ketemu,
        "laba_ditahan_pct": laba_ditahan_pct,
        "total_laba_bersih": round(total),
        "baris_matriks": baris_matriks,
        "nominal_matriks": nominal,
        "pencapaian_prioritas": pencapaian_prioritas,
        "pengali_pct": pengali, "label_pengali": label_pengali,
        "setelah_pengali": setelah_pengali,
        "goal": g["goal"], "potongan_pct": g["potongan_pct"],
        "rincian_potongan": g.get("rincian_potongan", []),
        "ambang_goal_pct": g.get("ambang_pct", 98),
        "potongan": potongan,
        "total": setelah_pengali - potongan,
    }
