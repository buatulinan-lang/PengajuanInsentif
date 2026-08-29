"""
Insentif Sales & Team.

Masukan: dua file ekspor mentah
  1. Data Pelanggan          -> ID Pelanggan, Nama Default Penjual, Kategori Pelanggan
  2. Rincian Faktur Penjualan -> tanggal, no faktur, id pelanggan, kategori barang,
                                 harga beli, total harga

Mekanisme (mengikuti tools Excel yang selama ini dipakai):
  - Tiap faktur diatribusikan ke sales lewat "Nama Default Penjual" milik
    pelanggannya, bukan ke orang yang menyerahkan faktur.
  - Omset sales = SERVICE (jasa + sparepart) + aksesoris + handphone + laptop + other.
    Omset inilah yang menentukan tarif (achievement).
  - Insentif aksesoris = 5% x gross profit aksesoris.
  - Insentif handphone / laptop = jumlah faktur x tarif sesuai tingkat achievement.
  - Insentif Team = 2% x bagi hasil service MFlash dari pelanggan Member Reguler.
Angka tarifnya ada di data/rules.json bagian "sales_team".
"""
import json, re, unicodedata
from collections import defaultdict
from openpyxl import load_workbook

from app.calc import RULES_PATH, CalcError, _num

KAT_SERVICE = ("JASA", "SPAREPART")
KAT_RETAIL = ("AKSESORIS", "HANDPHONE", "LAPTOP", "OTHER")


def _k(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    return re.sub(r"[^a-z0-9]", "", s.lower())


def _baca_tabel(path, wajib, batas_header=12):
    """Cari baris header yang memuat semua kolom `wajib`, kembalikan list-of-dict."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        rows.append(r)
        if i > 5000 and len(rows) > 5000:
            pass
    header_idx = None
    for i, r in enumerate(rows[:batas_header]):
        kunci = {_k(c) for c in r if c is not None}
        if all(_k(w) in kunci for w in wajib):
            header_idx = i
            break
    if header_idx is None:
        raise CalcError("Kolom " + ", ".join(wajib) + " tidak ditemukan di file. "
                        "Pastikan file yang diunggah benar.")
    header = [_k(c) for c in rows[header_idx]]
    out = []
    for r in rows[header_idx + 1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append({h: v for h, v in zip(header, r) if h})
    wb.close()
    return out


def _bulan(v):
    try:
        return v.month
    except AttributeError:
        try:
            return int(str(v).split("-")[1])
        except (IndexError, ValueError):
            return None


def _tarif(rules, omset, jenis):
    """Tarif per unit handphone/laptop menurut tingkat achievement omset."""
    for t in rules["achievement"]:
        if omset >= t["omset_min"]:
            return t[jenis]
    return rules["achievement"][-1][jenis]


KAT_PENJUALAN = {"handphone": "PENJUALAN HP", "laptop": "PENJUALAN LAPTOP",
                 "aksesoris": "PENJUALAN AKSESORIS"}


def _peta_nama(daftar_sales):
    """Semua ejaan (nama utama + alias) -> nama utama, dicocokkan tanpa huruf besar."""
    peta = {}
    for s in daftar_sales or []:
        peta[_k(s["nama"])] = s["nama"]
        for al in (s.get("alias") or "").split(","):
            if al.strip():
                peta[_k(al)] = s["nama"]
    return peta


def hitung_sales(path_pelanggan, path_faktur, bulan, tahun=None,
                 daftar_sales=None, pct_bagi_hasil_teknisi=None):
    """
    Omset & gross profit  : per baris barang, diatribusikan ke Nama Default Penjual
                            milik pelanggan.
    Jumlah unit HP/laptop : per faktur menurut Kategori Penjualan, diatribusikan ke
                            kolom "Yang Menyerahkan/Menjual".
    Keduanya mengikuti tools Excel yang selama ini dipakai.
    """
    rules = json.load(open(RULES_PATH))["sales_team"]
    if pct_bagi_hasil_teknisi is None:
        pct_bagi_hasil_teknisi = rules["pct_bagi_hasil_teknisi"]

    pelanggan = _baca_tabel(path_pelanggan, ["ID Pelanggan", "Nama Default Penjual"])
    faktur = _baca_tabel(path_faktur, ["NO FAKTUR", "ID PELANGGAN", "KATEGORI BARANG",
                                       "TOTAL HARGA", "KATEGORI PENJUALAN"])

    peta = _peta_nama(daftar_sales)

    def cocokkan(nama):
        return peta.get(_k(nama), "") if peta else str(nama or "").strip()

    penjual = {}
    for p in pelanggan:
        pid = str(p.get("idpelanggan") or "").strip()
        if pid:
            penjual[pid] = str(p.get("namadefaultpenjual") or "").strip()

    omset = defaultdict(lambda: defaultdict(float))
    gp = defaultdict(lambda: defaultdict(float))
    unit = defaultdict(lambda: defaultdict(set))
    bagi_hasil_member = 0.0
    dipakai = 0
    tak_dikenal = defaultdict(float)

    for f in faktur:
        if _bulan(f.get("tglfaktur")) != bulan:
            continue
        if tahun:
            th = getattr(f.get("tglfaktur"), "year", None)
            if th and th != tahun:
                continue
        dipakai += 1
        pid = str(f.get("idpelanggan") or "").strip()
        kat = str(f.get("kategoribarang") or "").strip().upper()
        kat_jual = str(f.get("kategoripenjualan") or "").strip().upper()
        total = _num(f.get("totalharga"))
        beli = _num(f.get("hargabeli"))
        no = str(f.get("nofaktur") or "")

        if kat == "JASA" and _k(f.get("kategoripelanggan")) == _k("MEMBER REGULER"):
            bagi_hasil_member += total * (1 - pct_bagi_hasil_teknisi / 100)

        # --- omset & gross profit: lewat Nama Default Penjual pelanggan
        asli = penjual.get(pid, "")
        sales = cocokkan(asli)
        if asli and not sales:
            tak_dikenal[asli] += total
        if sales:
            omset[sales][kat] += total
            if kat in ("AKSESORIS", "HANDPHONE", "LAPTOP") and beli:
                gp[sales][kat] += total - beli

        # --- jumlah unit: lewat kolom Yang Menyerahkan/Menjual
        asli2 = str(f.get("yangmenyerahkanmenjualfakturpenjualan") or "").strip()
        penjual_faktur = cocokkan(asli2)
        if asli2 and not penjual_faktur:
            tak_dikenal.setdefault(asli2, 0.0)
        if penjual_faktur and no:
            for kunci, label in KAT_PENJUALAN.items():
                if kat_jual == label:
                    unit[penjual_faktur][kunci].add(no)
            if kat_jual.startswith("SERVICE"):
                unit[penjual_faktur]["service"].add(no)

    nama_status = {s["nama"]: s.get("status", "") for s in (daftar_sales or [])}
    urutan = [s["nama"] for s in (daftar_sales or [])] or sorted(omset)

    baris, total_insentif = [], 0.0
    for nama in urutan:
        o = omset.get(nama, {})
        service = o.get("JASA", 0) + o.get("SPAREPART", 0)
        omset_total = service + sum(o.get(k, 0) for k in KAT_RETAIL)
        n_hp = len(unit[nama]["handphone"])
        n_lt = len(unit[nama]["laptop"])
        tarif_hp = _tarif(rules, omset_total, "handphone")
        tarif_lt = _tarif(rules, omset_total, "laptop")
        ins_aks = gp[nama]["AKSESORIS"] * rules["pct_insentif_aksesoris"] / 100
        ins_hp, ins_lt = n_hp * tarif_hp, n_lt * tarif_lt
        subtotal = ins_aks + ins_hp + ins_lt
        total_insentif += subtotal
        baris.append({
            "nama": nama, "status": nama_status.get(nama, ""),
            "service": service, "jasa": o.get("JASA", 0),
            "sparepart": o.get("SPAREPART", 0), "aksesoris": o.get("AKSESORIS", 0),
            "handphone": o.get("HANDPHONE", 0), "laptop": o.get("LAPTOP", 0),
            "other": o.get("OTHER", 0), "omset_total": omset_total,
            "gp_aksesoris": round(gp[nama]["AKSESORIS"]),
            "n_service": len(unit[nama]["service"]),
            "n_aksesoris": len(unit[nama]["aksesoris"]),
            "n_handphone": n_hp, "n_laptop": n_lt,
            "tarif_handphone": tarif_hp, "tarif_laptop": tarif_lt,
            "insentif_aksesoris": round(ins_aks),
            "insentif_handphone": ins_hp, "insentif_laptop": ins_lt,
            "total": round(subtotal),
        })

    insentif_team = round(bagi_hasil_member * rules["pct_insentif_team"] / 100)
    return {
        "jenis": "sales_team", "bulan": bulan, "tahun": tahun, "baris": baris,
        "subtotal_sales": round(total_insentif),
        "omset_service_member": round(bagi_hasil_member),
        "pct_bagi_hasil_teknisi": pct_bagi_hasil_teknisi,
        "pct_insentif_team": rules["pct_insentif_team"],
        "insentif_team": insentif_team,
        "total": round(total_insentif) + insentif_team,
        "jumlah_faktur_diproses": dipakai,
        "nama_tak_dikenal": sorted(tak_dikenal),
    }
