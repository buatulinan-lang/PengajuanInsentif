"""
Mesin perhitungan Insentif Profit Store Leader.

Sumber data: file Laporan Laba/Rugi cabang (mis. KLENDER.xlsx).
Tiap bulan menempati 3 kolom: [rincian] [total] [prosentase].
JANUARI = C/D/E, FEBRUARI = F/G/H, ... DESEMBER = AJ/AK/AL.

Alur:
  Laba Bersih  = Total Laba Bersih Setelah Prepaid - Laba Ditahan
  Persentase   = Laba Bersih / Jumlah Omzet   (penentu kolom matriks)
  Insentif     = matriks[baris laba terdekat ke bawah][kolom GP%]
Aturan angkanya ada di data/rules.json agar bisa diubah tanpa menyentuh kode.
"""
import json, os, re
from openpyxl import load_workbook

from app.paths import RULES_PATH

BULAN_XLS = ["JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI",
             "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]

# Label baris yang dicari di kolom B (dicocokkan longgar, tanpa spasi/huruf besar)
BARIS = {
    "omzet":            "jumlah omzet madinah flash",
    "laba_kotor":       "laba kotor sesudah fee teknisi",
    "fee_teknisi":      "total biaya fee teknisi",
    "biaya_operasional": "total biaya operasional",
    "laba_setelah_pajak": "total laba bersih setelah pajak",
    "laba_setelah_prepaid": "total laba bersih setelah prepaid",
}


class CalcError(Exception):
    pass


def load_rules():
    with open(RULES_PATH) as f:
        return json.load(f)


def _key(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def _num(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def baca_sheet(path, sheet=None):
    """Baca satu sheet menjadi daftar baris. Dibuka sekali, dipakai berulang."""
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        return [r for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _cari_kolom_bulan(rows, bulan):
    """Kembalikan (kolom_rincian, kolom_total, kolom_persen) untuk bulan ke-n.

    Nomor kolom dihitung mulai 1 agar sama dengan penomoran Excel.
    """
    target = _key(BULAN_XLS[bulan - 1])
    for r in rows[:40]:
        for i, v in enumerate(r):
            if _key(v) == target:
                return i + 1, i + 2, i + 3
    raise CalcError(f"Kolom bulan {BULAN_XLS[bulan - 1]} tidak ditemukan "
                    f"di file Excel.")


def _cari_baris(rows):
    """Peta nama internal -> indeks baris, dari label di kolom B."""
    hasil, dicari = {}, {k: _key(v) for k, v in BARIS.items()}
    for i, r in enumerate(rows):
        k = _key(r[1]) if len(r) > 1 else ""
        if not k:
            continue
        for nama, pola in dicari.items():
            if nama not in hasil and k == pola:
                hasil[nama] = i
    kurang = [BARIS[n] for n in BARIS if n not in hasil]
    if kurang:
        raise CalcError("Baris berikut tidak ditemukan di Excel: " + "; ".join(kurang))
    return hasil


def hitung_profit_sl(path, bulan, laba_ditahan_pct=None, sheet=None, rows=None):
    rules = load_rules()["profit_sl"]
    if laba_ditahan_pct is None:
        laba_ditahan_pct = rules.get("laba_ditahan_pct_default", 7)

    if rows is None:
        rows = baca_sheet(path, sheet)
    c_rinci, c_total, _ = _cari_kolom_bulan(rows, bulan)
    baris = _cari_baris(rows)

    def total(nama):
        r = rows[baris[nama]]
        return _num(r[c_total - 1]) if len(r) >= c_total else 0.0

    omzet          = total("omzet")
    laba_kotor     = total("laba_kotor")          # sesudah fee teknisi
    fee_teknisi    = total("fee_teknisi")
    biaya_ops      = total("biaya_operasional")
    setelah_pajak  = total("laba_setelah_pajak")
    setelah_prepaid = total("laba_setelah_prepaid")

    if omzet == 0:
        raise CalcError(f"Data bulan {BULAN_XLS[bulan-1].title()} masih kosong "
                        f"(omzet 0). Pilih bulan lain atau lengkapi Excel-nya.")

    prepaid = setelah_pajak - setelah_prepaid       # total PRE PAID EXPENSE
    laba_kotor_sebelum_fee = laba_kotor + fee_teknisi

    laba_ditahan = round(setelah_prepaid * laba_ditahan_pct / 100)
    laba_bersih = setelah_prepaid - laba_ditahan

    # penentu kolom matriks: laba bersih berbanding jumlah omzet
    gp_pct = laba_bersih / omzet * 100 if omzet else 0

    # kolom matriks: ambang tertinggi yang masih terpenuhi
    kolom = rules["kolom_gp"]
    idx_kolom, label_kolom = 0, kolom[0]["label"]
    for i, k in enumerate(kolom):
        if gp_pct >= k["min_pct"]:
            idx_kolom, label_kolom = i, k["label"]

    # baris matriks: turun ke ambang terdekat di bawah laba bersih
    insentif, baris_matriks = 0, None
    for m in rules["matriks"]:
        if laba_bersih >= m["laba_min"]:
            insentif, baris_matriks = m["insentif"][idx_kolom], m["laba_min"]

    return {
        "jenis": "profit_sl",
        "bulan": bulan,
        "bulan_nama": BULAN_XLS[bulan - 1].title(),
        "omzet": omzet,
        "laba_kotor_sesudah_fee": laba_kotor,
        "fee_teknisi": fee_teknisi,
        "laba_kotor_sebelum_fee": laba_kotor_sebelum_fee,
        "biaya_operasional": biaya_ops,
        "prepaid": prepaid,
        "laba_setelah_prepaid": setelah_prepaid,
        "laba_ditahan_pct": laba_ditahan_pct,
        "laba_ditahan": laba_ditahan,
        "laba_bersih": laba_bersih,
        "gp_pct": round(gp_pct, 2),
        "kolom_matriks": label_kolom,
        "baris_matriks": baris_matriks,
        "total": insentif,
    }


def compute(jenis, path, bulan=None, laba_ditahan_pct=None):
    if jenis == "profit_sl":
        return hitung_profit_sl(path, bulan, laba_ditahan_pct)
    raise CalcError(f"Perhitungan otomatis untuk '{jenis}' belum tersedia. "
                    f"Isi nilainya secara manual.")


# ---------------------------------------------------------------- rotasi
def pengali_rotasi(offset, tujuan_ge_asal):
    """Proporsi pengali (asal, tujuan) pada bulan ke-`offset` sejak mutasi."""
    rot = load_rules()["profit_sl"]["rotasi"]
    tabel = rot["tujuan_lebih_besar_atau_sama"] if tujuan_ge_asal else rot["tujuan_lebih_kecil"]
    if offset < 0:                      # bulan sebelum mutasi efektif
        return 100.0, 0.0
    i = min(offset, len(tabel["asal"]) - 1)
    return float(tabel["asal"][i]), float(tabel["tujuan"][i])


def label_bulan_rotasi(offset):
    if offset < 0:
        return "sebelum mutasi"
    return "Bulan H" if offset == 0 else f"Bulan H+{offset}"


def hitung_blok(blok, bulan, tahun, laba_ditahan_pct):
    """
    blok = {
      "tipe": "tunggal" | "rotasi",
      "asal":   {"cabang": str, "path": str},
      "tujuan": {"cabang": str, "path": str},   # hanya untuk rotasi
      "mutasi_bulan": int, "mutasi_tahun": int  # hanya untuk rotasi
    }
    """
    hasil_asal = hitung_profit_sl(blok["asal"]["path"], bulan, laba_ditahan_pct)

    if blok["tipe"] != "rotasi":
        return {
            "tipe": "tunggal",
            "asal": {"cabang": blok["asal"]["cabang"], "hasil": hasil_asal,
                     "pengali": 100.0, "insentif": hasil_asal["total"]},
            "total": hasil_asal["total"],
        }

    hasil_tujuan = hitung_profit_sl(blok["tujuan"]["path"], bulan, laba_ditahan_pct)
    offset = ((tahun * 12 + bulan) -
              (blok["mutasi_tahun"] * 12 + blok["mutasi_bulan"]))
    tujuan_ge = hasil_tujuan["laba_bersih"] >= hasil_asal["laba_bersih"]
    p_asal, p_tujuan = pengali_rotasi(offset, tujuan_ge)

    ins_asal = round(hasil_asal["total"] * p_asal / 100)
    ins_tujuan = round(hasil_tujuan["total"] * p_tujuan / 100)
    return {
        "tipe": "rotasi",
        "offset": offset,
        "offset_label": label_bulan_rotasi(offset),
        "kasus": ("Net Profit cabang tujuan lebih besar atau sama"
                  if tujuan_ge else "Net Profit cabang tujuan lebih kecil"),
        "mutasi": f"{BULAN_XLS[blok['mutasi_bulan']-1].title()} {blok['mutasi_tahun']}",
        "asal": {"cabang": blok["asal"]["cabang"], "hasil": hasil_asal,
                 "pengali": p_asal, "insentif": ins_asal},
        "tujuan": {"cabang": blok["tujuan"]["cabang"], "hasil": hasil_tujuan,
                   "pengali": p_tujuan, "insentif": ins_tujuan},
        "total": ins_asal + ins_tujuan,
    }


def hitung_pengurang_goal(goals):
    """Potongan proporsional bila pencapaian goal tidak memenuhi ambang.

    goals[0] adalah Goal Utama, sisanya goal pendamping. Aturannya:
      - Goal Utama tidak tercapai            -> potongan 10%
      - 1 goal pendamping tidak tercapai     -> tambah potongan 5%
      - lebih dari 1 tidak tercapai          -> tambah potongan 10%
    """
    ru = load_rules()["profit_sl"]["pengurang_goal"]
    ambang = ru["ambang_tercapai_pct"]
    rinci, pct = [], 0.0

    for i, g in enumerate(goals or []):
        nilai = g.get("pencapaian")
        tercapai = nilai is not None and float(nilai) >= ambang
        rinci.append({"nama": g.get("nama") or f"Goal {i+1}",
                      "pencapaian": None if nilai is None else round(float(nilai), 2),
                      "utama": i == 0, "tercapai": tercapai,
                      "diisi": nilai is not None})

    diisi = [g for g in rinci if g["diisi"]]
    if not diisi:
        return {"goal": rinci, "potongan_pct": 0.0, "rincian_potongan": []}

    potongan = []
    utama = rinci[0] if rinci else None
    if utama and utama["diisi"] and not utama["tercapai"]:
        pct += ru["potongan_goal_utama_pct"]
        potongan.append({"nama": f"Goal Utama ({utama['nama']}) tidak tercapai",
                         "pct": ru["potongan_goal_utama_pct"]})

    meleset = [g for g in rinci[1:] if g["diisi"] and not g["tercapai"]]
    if len(meleset) == 1:
        pct += ru["potongan_satu_pendamping_pct"]
        potongan.append({"nama": f"1 goal pendamping tidak tercapai "
                                 f"({meleset[0]['nama']})",
                         "pct": ru["potongan_satu_pendamping_pct"]})
    elif len(meleset) > 1:
        pct += ru["potongan_lebih_satu_pendamping_pct"]
        potongan.append({"nama": f"{len(meleset)} goal pendamping tidak tercapai "
                                 f"({', '.join(g['nama'] for g in meleset)})",
                         "pct": ru["potongan_lebih_satu_pendamping_pct"]})

    return {"goal": rinci, "potongan_pct": pct, "rincian_potongan": potongan,
            "ambang_pct": ambang}


def insentif_shift(kode):
    """Insentif shift masuk Store Leader: 3 hari/pekan atau 6 hari/pekan."""
    if not kode:
        return None
    rules = json.load(open(RULES_PATH)).get("shift_sl", {})
    for p in rules.get("pilihan", []):
        if str(p["kode"]) == str(kode):
            return dict(p)
    return None


def hitung_pengajuan(blok_list, bulan, tahun, laba_ditahan_pct=None, goals=None,
                     shift=None):
    """Hitung seluruh blok cabang pada satu pengajuan, lalu terapkan pengurang goal.

    Insentif shift masuk ditambahkan setelah potongan goal, karena nilainya
    tetap dan tidak bergantung pada pencapaian laba.
    """
    hasil, catatan = [], []
    for i, blok in enumerate(blok_list, 1):
        try:
            b = hitung_blok(blok, bulan, tahun, laba_ditahan_pct)
            b["nomor"] = i
            hasil.append(b)
        except CalcError as e:
            catatan.append(f"Blok {i}: {e}")

    # Satu cabang hanya boleh dihitung sekali. Store Leader yang berpindah
    # lewat beberapa cabang membuat satu cabang muncul di lebih dari satu blok
    # (mis. cabang tujuan blok pertama menjadi cabang asal blok berikutnya,
    # atau beberapa blok berbagi cabang tujuan yang sama). Tanpa penyaringan
    # ini cabang tersebut ikut terhitung berkali-kali.
    sisi = []
    for b in hasil:
        for peran in ("asal", "tujuan"):
            if b.get(peran):
                sisi.append((b, peran, b[peran]))
    dipakai, duplikat = {}, []
    # Yang dipertahankan adalah kemunculan dengan insentif terbesar, yaitu
    # pengali proporsional tertinggi untuk cabang itu.
    for b, peran, d in sorted(sisi, key=lambda x: -x[2]["insentif"]):
        kunci = re.sub(r"[^a-z0-9]", "", str(d["cabang"]).lower())
        if kunci in dipakai:
            d["dihitung"] = False
            d["alasan"] = f"sudah dihitung pada {dipakai[kunci]}"
            duplikat.append(d["cabang"])
        else:
            d["dihitung"] = True
            dipakai[kunci] = (f"blok {b['nomor']} sebagai cabang {peran}"
                              if len(hasil) > 1 else f"cabang {peran}")
    for b in hasil:
        b["total"] = sum(b[p]["insentif"] for p in ("asal", "tujuan")
                         if b.get(p) and b[p].get("dihitung"))

    subtotal = sum(b["total"] for b in hasil)
    if duplikat:
        catatan.append("Cabang berikut muncul lebih dari sekali dan hanya "
                       "dihitung satu kali: " + ", ".join(sorted(set(duplikat))))
    g = hitung_pengurang_goal(goals)
    potongan = round(subtotal * g["potongan_pct"] / 100)
    for p in g.get("rincian_potongan", []):
        p["nilai"] = round(subtotal * p["pct"] / 100)

    sh = insentif_shift(shift)
    nilai_shift = sh["nominal"] if sh else 0

    return {"blok": hasil, "subtotal_blok": subtotal,
            "goal": g["goal"], "potongan_pct": g["potongan_pct"],
            "rincian_potongan": g.get("rincian_potongan", []),
            "ambang_goal_pct": g.get("ambang_pct", 98),
            "potongan": potongan,
            "cabang_duplikat": sorted(set(duplikat)),
            "insentif_profit": subtotal - potongan,
            "shift": sh, "insentif_shift": nilai_shift,
            "total": subtotal - potongan + nilai_shift,
            "catatan": catatan}
